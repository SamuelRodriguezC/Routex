import openpyxl
from decimal import Decimal, InvalidOperation
from django.db import transaction, IntegrityError
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from logistics.models import Route, Status, Priority
from .execution_log_service import ExecutionLogService


def parse_aware_datetime(value):
    if not value:
        return None

    dt = parse_datetime(str(value))
    if dt and timezone.is_naive(dt):
        dt = timezone.make_aware(dt)
    return dt


class RouteImportService:

    REQUIRED_FIELDS = [
        "origin",
        "destination",
        "distance_km",
        "priority",
        "time_window_start",
        "time_window_end",
        "status",
    ]

    @staticmethod
    def import_routes(file):

        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        total_rows = sheet.max_row - 1
        processed = 0

        failed_status = Status.objects.get(description__iexact="FAILED")
        default_priority = Priority.objects.first()

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):

            processed += 1
            data = dict(zip(headers, row))
            row_errors = []

            # ==========================
            # Validaciones
            # ==========================

            for field in RouteImportService.REQUIRED_FIELDS:
                if not data.get(field):
                    row_errors.append(f"{field} es obligatorio")

            origin = str(data.get("origin", "")).strip()
            destination = str(data.get("destination", "")).strip()

            try:
                distance = Decimal(str(data.get("distance_km")))
                if distance <= 0:
                    row_errors.append("distance_km debe ser mayor que 0")
            except (InvalidOperation, TypeError):
                row_errors.append("distance_km inválido")
                distance = Decimal("0")

            start_date = parse_aware_datetime(data.get("time_window_start"))
            end_date = parse_aware_datetime(data.get("time_window_end"))

            if not start_date or not end_date:
                row_errors.append("Formato fecha inválido")
            elif start_date >= end_date:
                row_errors.append("Ventana de tiempo inválida")

            status_obj = Status.objects.filter(
                description__iexact=str(data.get("status", "")).strip()
            ).first()

            if not status_obj:
                row_errors.append("status no existe")

            priority_obj = Priority.objects.filter(
                id=data.get("priority")
            ).first()

            if not priority_obj:
                row_errors.append("priority no existe")

            # ==========================
            # Atomic POR FILA
            # ==========================

            try:
                with transaction.atomic():

                    if row_errors:

                        route = Route.objects.create(
                            origin=origin or "INVALID",
                            destination=destination or "INVALID",
                            distance_km=distance,
                            time_window_start=start_date,
                            time_window_end=end_date,
                            status=failed_status,
                            priority=priority_obj if priority_obj else default_priority
                        )

                        ExecutionLogService.log_multiple_errors(route, row_errors)

                    else:

                        route = Route.objects.create(
                            origin=origin,
                            destination=destination,
                            distance_km=distance,
                            time_window_start=start_date,
                            time_window_end=end_date,
                            status=status_obj,
                            priority=priority_obj,
                        )

                        ExecutionLogService.log(
                            route,
                            result="SUCCESS",
                            message="Ruta importada correctamente"
                        )

            except IntegrityError:

                existing_route = Route.objects.filter(
                    origin=origin,
                    destination=destination,
                    time_window_start=start_date,
                    time_window_end=end_date,
                ).first()

                if existing_route:
                    ExecutionLogService.log(
                        existing_route,
                        result="FAILED",
                        message="Intento de insertar ruta duplicada"
                    )

                continue

        return {
            "total_rows": total_rows,
            "processed": processed,
            "message": "Importación finalizada correctamente"
        }